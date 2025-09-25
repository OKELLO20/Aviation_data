
import requests
import csv
import pandas as pd
import openpyxl
import time

url = "https://api.aviationstack.com/v1/timetable?iataCode=NBO&type=arrival&access_key=88379c2b4cd1652465f13029212c63a6"

headers = {
    'Accept': 'application/json',
    'Content-Type': 'application/json'
}

response = requests.request("GET", url,headers=headers,data={})
myjson = response.json()
ourdata=[]
csvheader = ['Arrival_time', 'airline_name']

for x in myjson['data']:
    listing = [x['arrival']['actualTime'], x['airline']['name']]
    ourdata.append(listing)
print('DAN DA DAN')

with open('Users/Phanice/Desktop/AviationData/quil.csv','w', encoding='UTF8', newline='') as f:
    writer = csv.writer(f)

    writer.writerow(csvheader)
    writer.writerows(ourdata)

print('done')

csv= pd.read_csv('Users/Phanice/Desktop/AviationData/quil.csv')

excelWriter = pd.ExcelWriter('Users/Phanice/Desktop/AviationData/aviation.xlsx')
csv.to_excel(excelWriter)

excelWriter.close()
print('Dan Da Dan')

from openpyxl import load_workbook

wb = load_workbook('Users/Phanice/Desktop/AviationData/aviation.xlsx')
ws = wb.active

ws.insert_cols(2)
ws['B1'].value='City'

ws.insert_cols(3)
ws['C1'].value='Operation'


excel_file = 'Users/Phanice/Desktop/AviationData/aviation.xlsx'
df = pd.read_excel(excel_file, usecols=[0])

leg = df.max()

x = int(leg) + 3

for i in range(2,x):
    ws.cell(row=i, column=2).value = 'Nairobi'

wb.save('Users/Phanice/Desktop/AviationData/aviation.xlsx')


for i in range(2,x):
    ws.cell(row=i, column=3).value = 'Arrival'

wb.save('Users/Phanice/Desktop/AviationData/aviation.xlsx')

print('sleeping')

time.sleep(62)

print('working')

import requests
import csv
import pandas as pd
import os


url = "https://api.aviationstack.com/v1/timetable?iataCode=NBO&type=departure&access_key=88379c2b4cd1652465f13029212c63a6"

headers = {
    'Accept': 'application/json',
    'Content-Type': 'application/json'
}

response = requests.request("GET", url,headers=headers,data={})
ourjson = response.json()
ourdata=[]
csvheader = ['Departure_time', 'airline_name']

for x in ourjson['data']:
    listing = [x['departure']['actualTime'], x['airline']['name']]
    ourdata.append(listing)
print('DAN DA DAN')

with open('Users/Phanice/Desktop/AviationData/third.csv','w', encoding='UTF8', newline='') as f:
    writer = csv.writer(f)

    writer.writerow(csvheader)
    writer.writerows(ourdata)

print('done')

csv= pd.read_csv('Users/Phanice/Desktop/AviationData/third.csv')

excelWriter = pd.ExcelWriter('Users/Phanice/Desktop/AviationData/flying.xlsx')
csv.to_excel(excelWriter)

excelWriter.close()
print('Dan Da Dan')

from openpyxl import load_workbook

wb = load_workbook('Users/Phanice/Desktop/AviationData/flying.xlsx')
ws = wb.active

ws.insert_cols(2)
ws['B1'].value='City'

ws.insert_cols(3)
ws['C1'].value='Operation'


excel_file = 'Users/Phanice/Desktop/AviationData/flying.xlsx'
df = pd.read_excel(excel_file, usecols=[0])

leg = df.max()

x = int(leg) + 3

for i in range(2,x):
    ws.cell(row=i, column=2).value = 'Nairobi'

wb.save('Users/Phanice/Desktop/AviationData/flying.xlsx')


for i in range(2,x):
    ws.cell(row=i, column=3).value = 'Departure'

wb.save('Users/Phanice/Desktop/AviationData/flying.xlsx')

data_folder= '/Users/Phanice/Desktop/AviationData'

df = []

for file in os.listdir(data_folder):
    if file.endswith('xlsx'):
        print('loading file {0}...'.format(file))
        df.append(pd.read_excel(os.path.join(data_folder,file)))

df_master = pd.concat(df)
df_master.to_excel('master file.xlsx', index=False)

